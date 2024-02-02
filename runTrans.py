import torch
from transformers import AutoProcessor, SeamlessM4Tv2Model

from logging import info, warning, error, basicConfig, FileHandler, StreamHandler
from openpyxl import Workbook, load_workbook

class Translator:
    def __init__(self, modelPath):
        self.processor = AutoProcessor.from_pretrained(modelPath)
        self.model = SeamlessM4Tv2Model.from_pretrained(modelPath)

    def translate_text(self, text, src_lang, tgt_lang):
        text_inputs = self.processor(text=text, src_lang=src_lang, return_tensors="pt")
        output_tokens = self.model.generate(**text_inputs, tgt_lang=tgt_lang, text_num_beams=5, generate_speech=False)
        translated_text = self.processor.decode(output_tokens[0].tolist()[0], skip_special_tokens=True)

        return translated_text

class Excel2Excel:
    """ Excel file to Sqlite """
    @staticmethod
    def _get_excel_row(
            excel_sheet, row_start=None, row_end=None, col_start=None,
            col_end=None):
        ''' Strip cell and remove none values, excel_sheet is worksheet in openpyxl workbook '''

        table_strip = []
        for row in excel_sheet.iter_rows(
                min_row=row_start, max_row=row_end, min_col=col_start,
                max_col=col_end, values_only=True):
            # 'None' to ''
            row_strip_wo_none = []
            for i in row:
                if str(i) == 'None':
                    row_strip_wo_none.append('')
                else:
                    row_strip_wo_none.append(str(i).strip())
            table_strip.append(row_strip_wo_none)
            print("row_strip_wo_none: " + str(row_strip_wo_none))
        return table_strip

    @staticmethod
    def excel_to_excel(excel_file, excel_sheets=None):
        """ Convert excel file to sqlite file: excel_sheets is list (['sheetname1', 'sheetname2']) OR None for all sheets """
        wb = load_workbook(filename=excel_file, data_only=True)
#        if not db_name:
#            db_name = excel_file.split('.xl')[0]

        col_before = 3
        col_after = 4
        for sheet in wb:
            sheet_name = sheet.title
            # check excel_sheets specify
            if excel_sheets and sheet_name in excel_sheets:
                sheet = wb[sheet_name]
            
            data = Excel2Excel._get_excel_row(sheet, row_start=5, row_end=50, col_start=col_before, col_end=col_before)
            return data    

    @staticmethod
    def excel_run():
        """ run with sqljoin yaml file """
        excel_file = '20231130_132930.xlsx'
#        excel_sheets = None
        excel_sheets = 'Sheet1'

        table_strip = []
        table_strip_out = []
        table_strip = Excel2Excel.excel_to_excel(
            excel_file, excel_sheets=excel_sheets)
        info(f'[{excel_file}] EXCEL TO EXCEL COMPLETED!')
#            except Exception as e:
#                error(f'check excel @ {part} : {e}')
#                raise SystemExit

        return table_strip   
def main():
    ''' main function to run parse2excel '''
    src_lang = "eng"
    tgt_lang = "vie"
    #sentence = "Contemplations before chanting."
#    sentence = '''The same is true with feelings, perceptions,
#    mental formations and consciousness.
#    Listen, Sariputra,
#    all dharmas are marked with emptiness.
#    They are neither produced nor destroyed,
#    neither defiled nor immaculate,
#    neither increasing nor decreasing.
#    Therefore an emptiness is neither forms
#    '''
    sentence = " เพลงเก่าฮิต ฟังกี่ครั้งก็อิน ( เพลงยุค90/เพลงยุค2000) 【LONGPLAY】 "
    
    modelPath = "/Volumes/OS/facebook/seamless-m4t-v2-large"
    translator = Translator(modelPath)
    translated_sentence = translator.translate_text(sentence, src_lang, tgt_lang)
    print(translated_sentence)
    
    # translator = Translator(modelPath)
    table_strip_out = []
    # table_strip = Excel2Excel.excel_run()
#    for sentence in table_strip:
#        translated_sentence = translator.translate_text(sentence, src_lang, tgt_lang)
#        print(translated_sentence)
#        table_strip_out.append(translated_sentence)

if __name__ == "__main__":
    main()
